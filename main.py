import io, os, math
from typing import Optional, Dict, List
from fastapi import FastAPI, Header, HTTPException, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# -------------------- Auth & App --------------------
API_KEY = os.getenv("API_KEY", "CHANGE_ME")

app = FastAPI(title="SD Tax Engine", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_methods=["*"], allow_headers=["*"], allow_credentials=False
)

def check_key(x_api_key: Optional[str]):
    if x_api_key != API_KEY:
        raise HTTPException(status_code=401, detail="Invalid API key")

# Health / root
@app.get("/")
def root(): return {"status": "ok", "message": "Tax Draft Backend is running"}
@app.get("/ping")
def ping(): return {"message": "pong"}

# -------------------- Models (simple & flexible) --------------------
class Meta(BaseModel):
    business_name: str
    proprietor_name: str
    pan: str
    dob: str              # YYYY-MM-DD
    fy: str               # e.g. 2024-25

class YearIn(BaseModel):
    # Minimal inputs; everything else can be 0 / omitted
    turnover: float = 0
    other_income: float = 0
    opening_stock: float = 0
    purchases: float = 0
    return_inwards: float = 0
    return_outwards: float = 0
    carriage_inward: float = 0
    closing_stock: float = 0

    # expenses (indirect)
    salaries: float = 0
    rent_utilities: float = 0
    admin_misc: float = 0
    depreciation_it: float = 0
    finance_costs: float = 0
    other_expenses: float = 0

    # balance sheet heads (optional)
    capital_open: float = 0
    additional_investment: float = 0
    drawings: float = 0
    loans: float = 0
    payables: float = 0
    receivables: float = 0
    fixed_assets: float = 0
    cash_bank: float = 0
    other_assets: float = 0
    other_liab: float = 0

class ExcelRequest(BaseModel):
    meta: Meta
    y1: YearIn
    # If y2 not provided, set auto_growth=True to generate from y1
    y2: Optional[YearIn] = None
    auto_growth: bool = False
    growth_pct: float = Field(10, ge=0, le=100)  # default 10% (fits 8-12 band)
    enforce_continuity: bool = True              # stock & capital continuity

# -------------------- Core helpers --------------------
def round_inr(x: float) -> float:
    return float(round(x or 0))

def apply_growth(base: YearIn, g: float) -> YearIn:
    f = 1 + g/100.0
    return YearIn(
        turnover=base.turnover*f,
        other_income=base.other_income*f,
        opening_stock=base.opening_stock*f,
        purchases=base.purchases*f,
        return_inwards=base.return_inwards*f,
        return_outwards=base.return_outwards*f,
        carriage_inward=base.carriage_inward*f,
        closing_stock=base.closing_stock*f,
        salaries=base.salaries*f,
        rent_utilities=base.rent_utilities*f,
        admin_misc=base.admin_misc*f,
        depreciation_it=base.depreciation_it*f,
        finance_costs=base.finance_costs*f,
        other_expenses=base.other_expenses*f,
        capital_open=base.capital_open*f,
        additional_investment=base.additional_investment*f,
        drawings=base.drawings*f,
        loans=base.loans*f,
        payables=base.payables*f,
        receivables=base.receivables*f,
        fixed_assets=base.fixed_assets*f,
        cash_bank=base.cash_bank*f,
        other_assets=base.other_assets*f,
        other_liab=base.other_liab*f
    )

def compute_profit(y: YearIn) -> Dict[str, float]:
    # Trading section
    net_purchases = y.purchases - y.return_outwards
    net_sales = y.turnover - y.return_inwards
    cogs = y.opening_stock + net_purchases + y.carriage_inward - y.closing_stock
    gross_profit = net_sales - cogs

    indirect_exp = (y.salaries + y.rent_utilities + y.admin_misc +
                    y.depreciation_it + y.finance_costs + y.other_expenses)

    net_profit = gross_profit + y.other_income - indirect_exp
    return {
        "net_sales": net_sales, "cogs": cogs, "gross_profit": gross_profit,
        "indirect_exp": indirect_exp, "net_profit": net_profit
    }

def simple_tax(total_income: float) -> Dict[str, float]:
    # Very light new-regime demo; replace with exact slabs later
    TI = max(0.0, total_income)
    slabs = [(300000,0.0),(400000,0.05),(300000,0.10),(200000,0.15),(300000,0.20),(10**18,0.30)]
    t, rem = 0.0, TI
    for width, rate in slabs:
        amt = min(rem, width)
        t += amt*rate
        rem -= amt
        if rem <= 0: break
    cess = 0.04*t
    return {"tax": round_inr(t), "cess": round_inr(cess), "total": round_inr(t+cess)}

# -------------------- Excel rendering helpers --------------------
BOLD = Font(bold=True); HDR = Font(size=12, bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
FILL = PatternFill("solid", fgColor="F2F2F2")

def header_block(ws, meta: Meta, title: str):
    ws.append([title]); ws["A1"].font = HDR
    ws.append([f"Business Name: {meta.business_name}"])
    ws.append([f"Proprietor Name: {meta.proprietor_name}"])
    ws.append([f"PAN: {meta.pan}"])
    ws.append([f"Date of Birth: {meta.dob}"])
    ws.append([f"Financial Year: {meta.fy}"])
    ws.append([""])

def t_table(ws, left_rows: List[List], right_rows: List[List]):
    ws.append(["Particulars", "Amount (₹)", "Particulars", "Amount (₹)"])
    for c in range(1,5):
        ws.cell(ws.max_row, c).font = BOLD
        ws.cell(ws.max_row, c).alignment = CENTER
        ws.cell(ws.max_row, c).fill = FILL
        ws.cell(ws.max_row, c).border = THIN
    maxlen = max(len(left_rows), len(right_rows))
    for i in range(maxlen):
        L = left_rows[i] if i < len(left_rows) else ["", 0]
        R = right_rows[i] if i < len(right_rows) else ["", 0]
        ws.append([L[0], round_inr(L[1] or 0), R[0], round_inr(R[1] or 0)])
        for col in (2,4):
            cell = ws.cell(ws.max_row, col); cell.number_format = "#,##0"; cell.alignment = RIGHT; cell.border = THIN
        for col in (1,3):
            ws.cell(ws.max_row, col).border = THIN
    ws.append([""])

def pl_t_format(y: YearIn) -> Dict[str, List[List]]:
    numbers = compute_profit(y)
    dr = [
        ["To Opening Stock", y.opening_stock],
        ["To Purchases", y.purchases],
        ["(-) Return Outwards", -y.return_outwards],
        ["To Carriage Inward", y.carriage_inward],
        ["To Salaries", y.salaries],
        ["To Rent & Utilities", y.rent_utilities],
        ["To Admin & Misc", y.admin_misc],
        ["To Finance Costs", y.finance_costs],
        ["To Depreciation (IT)", y.depreciation_it],
        ["To Other Expenses", y.other_expenses],
    ]
    cr = [
        ["By Sales", y.turnover],
        ["(-) Return Inwards", -y.return_inwards],
        ["By Closing Stock", y.closing_stock],
        ["By Other Income", y.other_income],
    ]
    # Balance with Net Profit/Gross Profit
    total_dr = sum(x[1] for x in dr)
    total_cr = sum(x[1] for x in cr)
    if total_cr >= total_dr:
        dr.append(["To Net Profit", total_cr - total_dr])
    else:
        cr.append(["By Gross Profit", total_dr - total_cr])
    return {"dr": dr, "cr": cr, "net_profit": numbers["net_profit"]}

def capital_t_format(opening: float, profit: float, addl_invest: float, drawings: float) -> Dict[str, List[List]]:
    cr = [
        ["By Opening Balance", opening],
        ["By Profit", profit],
        ["By Additional Investment", addl_invest],
    ]
    dr = [["To Drawings", drawings]]
    tot_cr = sum(x[1] for x in cr)
    tot_dr = sum(x[1] for x in dr)
    closing = max(0.0, tot_cr - tot_dr)
    dr.append(["To Closing Balance", closing])
    return {"dr": dr, "cr": cr, "closing": closing}

def balance_sheet_t_format(capital: float, y: YearIn) -> Dict[str, List[List]]:
    liabilities = [
        ["Capital", capital],
        ["Loans", y.loans],
        ["Payables", y.payables],
        ["Other Liabilities", y.other_liab],
    ]
    assets = [
        ["Fixed Assets", y.fixed_assets],
        ["Inventory", y.closing_stock],
        ["Receivables", y.receivables],
        ["Cash/Bank", y.cash_bank],
        ["Other Assets", y.other_assets],
    ]
    # Balance the sheet
    tot_l = sum(x[1] for x in liabilities)
    tot_a = sum(x[1] for x in assets)
    if tot_l > tot_a:
        assets.append(["Balancing Figure (Assets)", tot_l - tot_a])
    elif tot_a > tot_l:
        liabilities.append(["Balancing Figure (Liabilities)", tot_a - tot_l])
    return {"liabilities": liabilities, "assets": assets}

# -------------------- /excel real report --------------------
@app.post("/excel")
def generate_excel(payload: ExcelRequest, x_api_key: Optional[str] = Header(None)):
    check_key(x_api_key)

    meta = payload.meta
    y1 = payload.y1
    notes: List[str] = []

    # Year-2 handling
    y2 = payload.y2
    growth_used = None
    if not y2 and payload.auto_growth:
        y2 = apply_growth(y1, payload.growth_pct)
        growth_used = payload.growth_pct
        notes.append(f"Year-2 auto-generated at +{payload.growth_pct}% growth on all heads.")

    # Continuity rules (if both years present)
    if payload.enforce_continuity and y2:
        # Capital continuity requires y1 closing; compute after PL/Capital
        notes.append("Continuity applied: Closing Stock (Y1) = Opening Stock (Y2).")
        y2.opening_stock = y1.closing_stock

    # Start workbook
    wb = Workbook()

    # ---------- Sheet 1: Trading & Profit and Loss (T-format) ----------
    ws1 = wb.active; ws1.title = "Profit & Loss"
    header_block(ws1, meta, f"Trading and Profit & Loss Account")
    # Y1 block
    ws1.append(["Year-1"]); ws1["A"+str(ws1.max_row)].font = BOLD
    pl1 = pl_t_format(y1)
    t_table(ws1, pl1["dr"], pl1["cr"])
    np1 = round_inr(pl1["net_profit"])

    # Y2 block (optional)
    cap1 = capital_t_format(y1.capital_open, np1, y1.additional_investment, y1.drawings)
    cap1_close = cap1["closing"]

    if y2:
        ws1.append(["Year-2" + (f" (Adjusted @ {growth_used}%)" if growth_used else "")]); ws1["A"+str(ws1.max_row)].font = BOLD
        pl2 = pl_t_format(y2)
        t_table(ws1, pl2["dr"], pl2["cr"])
        np2 = round_inr(pl2["net_profit"])
        cap2 = capital_t_format(
            opening=cap1_close if payload.enforce_continuity else y2.capital_open,
            profit=np2, addl_invest=y2.additional_investment, drawings=y2.drawings
        )
        cap2_close = cap2["closing"]
    else:
        np2 = cap2_close = None

    # ---------- Sheet 2: Balance Sheet then Capital Account ----------
    ws2 = wb.create_sheet("Balance Sheet & Capital")
    header_block(ws2, meta, "Balance Sheet")

    def render_bs_cap_block(label: str, y: YearIn, opening_capital: float, profit_value: float):
        ws2.append([label]); ws2["A"+str(ws2.max_row)].font = BOLD
        cap_stmt = capital_t_format(opening_capital, profit_value, y.additional_investment, y.drawings)
        bs = balance_sheet_t_format(cap_stmt["closing"], y)
        # Balance Sheet (T)
        t_table(ws2, bs["liabilities"], bs["assets"])
        # Capital Account (T)
        ws2.append(["Proprietor's Capital Account"]); ws2["A"+str(ws2.max_row)].font = BOLD
        t_table(ws2, cap_stmt["dr"], cap_stmt["cr"])
        return cap_stmt

    # Y1
    render_bs_cap_block("Year-1", y1, y1.capital_open, np1)

    # Y2
    if y2:
        opening_cap_y2 = cap1_close if payload.enforce_continuity else y2.capital_open
        cap_stmt_y2 = render_bs_cap_block(
            "Year-2" + (f" (Adjusted @ {growth_used}%)" if growth_used else ""),
            y2, opening_cap_y2, round_inr(pl2["net_profit"])
        )
        if payload.enforce_continuity:
            # Note continuity checks
            notes.append(f"Capital continuity: Year-1 closing capital carried as Year-2 opening capital ({round_inr(opening_cap_y2)}).")

    # ---------- Sheet 3: Computation of Income (vertical) ----------
    ws3 = wb.create_sheet("Computation of Income")
    header_block(ws3, meta, "Computation of Total Income (Demo)")

    def comp_block(title: str, profit: float):
        ws3.append([title]); ws3["A"+str(ws3.max_row)].font = BOLD
        tax = simple_tax(profit if profit is not None else 0)
        rows = [
            ("Income from Business or Profession", profit or 0),
            ("Add: Other Income", 0),
            ("Gross Total Income", profit or 0),
            ("Less: Chapter VI-A", 0),
            ("Total Income", profit or 0),
            ("Tax on Total Income", tax["tax"]),
            ("Add: Health & Education Cess", tax["cess"]),
            ("Total Tax Liability", tax["total"]),
            ("Less: Taxes Paid (TDS/TCS/Adv/SAT)", 0),
            ("Net Tax Payable/(Refund)", tax["total"]),
        ]
        ws3.append(["Particulars", "Amount (₹)"])
        for c in (1,2):
            ws3.cell(ws3.max_row, c).font = BOLD
            ws3.cell(ws3.max_row, c).alignment = CENTER
            ws3.cell(ws3.max_row, c).fill = FILL
            ws3.cell(ws3.max_row, c).border = THIN
        for k, v in rows:
            ws3.append([k, round_inr(v)])
            ws3.cell(ws3.max_row, 1).border = THIN
            cell = ws3.cell(ws3.max_row, 2); cell.border = THIN; cell.alignment = RIGHT; cell.number_format = "#,##0"
        ws3.append([""])

    comp_block("Year-1", np1)
    if y2: comp_block("Year-2" + (f" (Adjusted @ {growth_used}%)" if growth_used else ""), np2)

    # ---------- Notes sheet ----------
    notes_ws = wb.create_sheet("Notes")
    notes_ws["A1"] = "Notes"; notes_ws["A1"].font = HDR
    r = 3
    for n in notes:
        notes_ws[f"A{r}"] = f"• {n}"; r += 1

    # ---------- Return workbook ----------
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename=\"Tax_Report.xlsx\"'}
    )
